"""手表指令历史导出为 Excel（客户可读中文列，不含原始帧/JSON）。"""

from __future__ import annotations

import json
import re
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from app.db.models import CommandEvent
from app.web.humanize import data_type_label, summary_from_parsed
from app.web.timefmt import format_local_time


def _attachment_note(ev: CommandEvent) -> str:
    if not ev.media_path:
        return "无"
    name = Path(ev.media_path).name
    cmd = (ev.command or "").upper()
    if cmd == "SENDPHOTO":
        kind = "照片"
    elif cmd == "JXTK":
        kind = "语音"
    else:
        kind = "文件"
    return f"有（{kind}），文件名：{name}，请在管理后台「手表详情」中下载"


def build_device_history_xlsx(events: list[CommandEvent]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "历史记录"

    headers = ["序号", "上报时间", "数据类型", "内容说明", "附件"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    wrap = Alignment(wrap_text=True, vertical="top")
    for i, ev in enumerate(events, start=1):
        try:
            sj = json.loads(ev.summary_json) if ev.summary_json else {}
        except json.JSONDecodeError:
            sj = {}
        if not isinstance(sj, dict):
            sj = {}
        summary = summary_from_parsed(ev.command, sj)
        row = [
            i,
            format_local_time(ev.created_at),
            data_type_label(ev.command),
            summary,
            _attachment_note(ev),
        ]
        ws.append(row)
        for c in range(1, 6):
            ws.cell(row=i + 1, column=c).alignment = wrap

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 48
    ws.column_dimensions["E"].width = 36

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def safe_filename_part(device_id: str) -> str:
    s = re.sub(r"[^\w\-.]", "_", device_id).strip("_") or "device"
    return s[:48]
